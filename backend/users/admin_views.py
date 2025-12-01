"""
用户模块 - 管理员接口

包含：用户CRUD、批量操作、导入导出、激活码管理、学生画像管理
"""
import codecs
import csv
import io
import logging
import secrets
import string
from collections.abc import Mapping
from datetime import timedelta
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from django.utils import timezone
from django.db import transaction, models
from django.http import HttpResponse

from common.responses import success_response, created_response, error_response
from common.permissions import IsAdmin
from courses.models import Enrollment
from knowledge.models import KnowledgeMastery
from assessments.models import AbilityScore
from .models import User, HabitPreference, ActivationCode
from .serializers import HabitPreferenceSerializer


logger = logging.getLogger(__name__)
# UTF-8 BOM 文本，避免直接写字面量触发拼写告警。
UTF8_BOM = codecs.BOM_UTF8.decode('utf-8')


def _parse_pagination(
    query_params: Mapping[str, str],
    size_key: str = 'page_size',
    default_size: int = 20,
    max_size: int = 100,
) -> tuple[int, int]:
    """
    统一解析分页参数并兜底非法输入。
    :param query_params: 请求查询参数。
    :param size_key: 每页数量对应的参数名。
    :param default_size: 默认每页数量。
    :param max_size: 每页数量上限。
    :return: 规范化后的页码与每页数量。
    """

    try:
        page = max(1, int(query_params.get('page', 1)))
        page_size = min(max(1, int(query_params.get(size_key, default_size))), max_size)
    except (ValueError, TypeError):
        page = 1
        page_size = default_size

    return page, page_size


# ============ 激活码管理 ============

@api_view(['POST'])
@permission_classes([IsAuthenticated, IsAdmin])
def generate_activation_code(request):
    """
    生成激活码（仅管理员）
    POST /api/admin/activation-codes/generate
    
    请求参数：
    - code_type: 激活码类型，teacher/admin
    - count: 生成数量（默认1，最大50）
    - expires_days: 过期天数（可选，0表示永不过期）
    - remark: 备注（可选）
    """
    user = request.user

    code_type = request.data.get('code_type', 'teacher')
    if code_type not in ['teacher', 'admin']:
        return error_response(msg='无效的激活码类型', code=400)

    try:
        count = min(int(request.data.get('count', 1)), 50)
        if count < 1:
            count = 1
    except (ValueError, TypeError):
        count = 1

    try:
        expires_days = int(request.data.get('expires_days', 0))
        if expires_days < 0:
            expires_days = 0
    except (ValueError, TypeError):
        expires_days = 0

    remark = request.data.get('remark', '')

    expires_at = None
    if expires_days > 0:
        expires_at = timezone.now() + timedelta(days=expires_days)

    codes = []

    for _ in range(count):
        code = ActivationCode.objects.create(
            code=ActivationCode.generate_code(),
            code_type=code_type,
            created_by=user,
            expires_at=expires_at,
            remark=remark
        )
        codes.append({
            'code': code.code,
            'code_type': code.code_type,
            'expires_at': code.expires_at.isoformat() if code.expires_at else None
        })

    return created_response(
        data={'codes': codes},
        msg=f'成功生成 {count} 个激活码'
    )


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdmin])
def list_activation_codes(request):
    """
    获取激活码列表（仅管理员）
    GET /api/admin/activation-codes
    
    查询参数：
    - code_type: 筛选类型（可选）
    - is_used: 是否已使用（可选）
    - page: 页码
    - page_size: 每页数量
    """
    queryset = ActivationCode.objects.all()

    code_type = request.query_params.get('code_type')
    if code_type:
        queryset = queryset.filter(code_type=code_type)

    is_used = request.query_params.get('is_used')
    if is_used is not None:
        queryset = queryset.filter(is_used=is_used.lower() == 'true')

    page, page_size = _parse_pagination(request.query_params)
    start = (page - 1) * page_size
    end = start + page_size

    total = queryset.count()
    codes = queryset.select_related('used_by')[start:end]

    return success_response(
        data={
            'total': total,
            'page': page,
            'page_size': page_size,
            'codes': [
                {
                    'id': c.id,
                    'code': c.code,
                    'code_type': c.code_type,
                    'is_used': c.is_used,
                    'used_by': c.used_by.username if c.used_by else None,
                    'used_at': c.used_at.isoformat() if c.used_at else None,
                    'expires_at': c.expires_at.isoformat() if c.expires_at else None,
                    'remark': c.remark,
                    'created_at': c.created_at.isoformat()
                }
                for c in codes
            ]
        }
    )


@api_view(['GET', 'DELETE'])
@permission_classes([IsAuthenticated, IsAdmin])
def delete_activation_code(request, code_id):
    """
    获取激活码详情 / 删除激活码（仅管理员）
    GET /api/admin/activation-codes/{code_id}
    DELETE /api/admin/activation-codes/{code_id}
    """
    try:
        code = ActivationCode.objects.get(id=code_id)
    except ActivationCode.DoesNotExist:
        return error_response(msg='激活码不存在', code=404)

    if request.method == 'GET':
        return success_response(data={
            'id': code.id,
            'code': code.code,
            'code_type': code.code_type,
            'is_used': code.is_used,
            'used_by': code.used_by.username if code.used_by else None,
            'used_at': code.used_at.isoformat() if code.used_at else None,
            'expires_at': code.expires_at.isoformat() if code.expires_at else None,
            'remark': code.remark,
            'created_at': code.created_at.isoformat(),
        })

    if code.is_used:
        return error_response(msg='已使用的激活码不能删除', code=400)

    code.delete()
    return success_response(msg='激活码已删除')


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdmin])
def activation_code_detail(request, code_id):
    """
    获取激活码详情（管理员）
    GET /api/admin/activation-codes/{code_id}
    """
    logger.debug('管理员查看激活码详情: admin=%s code_id=%s', getattr(request.user, 'id', None), code_id)

    try:
        c = ActivationCode.objects.get(id=code_id)
    except ActivationCode.DoesNotExist:
        return error_response(msg='激活码不存在', code=404)

    return success_response(data={
        'id': c.id,
        'code': c.code,
        'code_type': c.code_type,
        'is_used': c.is_used,
        'used_by': c.used_by.username if c.used_by else None,
        'used_at': c.used_at.isoformat() if c.used_at else None,
        'expires_at': c.expires_at.isoformat() if c.expires_at else None,
        'remark': c.remark,
        'created_at': c.created_at.isoformat(),
        'created_by': c.created_by.username if c.created_by else None,
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsAdmin])
def activation_code_batch_delete(request):
    """
    批量删除激活码（管理员）
    POST /api/admin/activation-codes/batch-delete
    """

    code_ids = request.data.get('code_ids', [])
    if not code_ids:
        return error_response(msg='请提供激活码ID列表', code=400)

    deleted_count, _ = ActivationCode.objects.filter(id__in=code_ids, is_used=False).delete()
    return success_response(data={'deleted_count': deleted_count}, msg=f'已删除 {deleted_count} 个激活码')


@api_view(['POST'])
@permission_classes([])
def activation_code_validate(request):
    """
    验证激活码有效性
    POST /api/admin/activation-codes/validate
    """
    code_str = request.data.get('code')
    if not code_str:
        return error_response(msg='请提供激活码', code=400)

    try:
        c = ActivationCode.objects.get(code=code_str)
    except ActivationCode.DoesNotExist:
        return error_response(msg='激活码不存在', code=404)

    valid = c.is_valid()
    return success_response(data={
        'code': c.code,
        'code_type': c.code_type,
        'is_valid': valid,
        'is_used': c.is_used,
        'expires_at': c.expires_at.isoformat() if c.expires_at else None,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdmin])
def activation_code_export(request):
    """
    导出激活码列表为 CSV
    GET /api/admin/activation-codes/export
    """

    queryset = ActivationCode.objects.all().select_related('used_by', 'created_by')
    code_type = request.query_params.get('code_type')
    if code_type:
        queryset = queryset.filter(code_type=code_type)

    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="activation_codes.csv"'
    response.write(UTF8_BOM)

    writer = csv.writer(response)
    writer.writerow(['ID', '激活码', '类型', '已使用', '使用者', '使用时间', '过期时间', '备注', '创建时间'])

    for c in queryset[:10000]:
        writer.writerow([
            c.id, c.code, c.code_type,
            '是' if c.is_used else '否',
            c.used_by.username if c.used_by else '',
            c.used_at.strftime('%Y-%m-%d %H:%M:%S') if c.used_at else '',
            c.expires_at.strftime('%Y-%m-%d %H:%M:%S') if c.expires_at else '永不过期',
            c.remark or '',
            c.created_at.strftime('%Y-%m-%d %H:%M:%S'),
        ])

    return response


# ============ 用户管理 ============

@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdmin])
def admin_user_list(request):
    """
    获取用户列表（管理员）
    GET /api/admin/users
    
    查询参数：
    - page: 页码（默认1）
    - size: 每页数量（默认20）
    - role: 角色筛选（student/teacher/admin）
    - query: 搜索关键词（用户名/邮箱/手机号）
    - status: 状态筛选（active/inactive）
    """
    page, size = _parse_pagination(request.query_params, size_key='size')
    role = request.query_params.get('role')
    query = request.query_params.get('query')
    status = request.query_params.get('status')

    users = User.objects.all().order_by('-date_joined')

    if role:
        users = users.filter(role=role)

    if query:
        users = users.filter(
            models.Q(username__icontains=query) |
            models.Q(email__icontains=query) |
            models.Q(phone__icontains=query) |
            models.Q(real_name__icontains=query)
        )

    if status == 'active':
        users = users.filter(is_active=True)
    elif status == 'inactive':
        users = users.filter(is_active=False)

    total = users.count()
    start = (page - 1) * size
    end = start + size
    users = users[start:end]

    user_list = [{
        'user_id': u.id,
        'username': u.username,
        'email': u.email or '',
        'phone': u.phone or '',
        'real_name': u.real_name or '',
        'role': u.role,
        'is_active': u.is_active,
        'date_joined': u.date_joined.isoformat()
    } for u in users]

    return success_response(data={
        'total': total,
        'page': page,
        'size': size,
        'users': user_list
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdmin])
def admin_user_detail(request, user_id):
    """
    获取用户详情（管理员）
    GET /api/admin/users/{user_id}
    """
    try:
        target_user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return error_response(msg='用户不存在', code=404)

    return success_response(data={
        'user_id': target_user.id,
        'username': target_user.username,
        'email': target_user.email or '',
        'phone': target_user.phone or '',
        'real_name': target_user.real_name or '',
        'student_id': target_user.student_id or '',
        'role': target_user.role,
        'is_active': target_user.is_active,
        'avatar': target_user.avatar.url if target_user.avatar else None,
        'date_joined': target_user.date_joined.isoformat(),
        'last_login': target_user.last_login.isoformat() if target_user.last_login else None
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsAdmin])
def admin_user_create(request):
    """
    创建用户（管理员）
    POST /api/admin/users/create
    
    请求参数：
    - username: 用户名（必填）
    - password: 密码（必填）
    - role: 角色（必填，student/teacher/admin）
    - email: 邮箱（选填）
    - phone: 手机号（选填）
    - real_name: 真实姓名（选填）
    """
    username = request.data.get('username')
    password = request.data.get('password')
    role = request.data.get('role', 'student')
    email = request.data.get('email', '')
    phone = request.data.get('phone', '')
    real_name = request.data.get('real_name', '')

    if not username or not password:
        return error_response(msg='用户名和密码不能为空', code=400)

    if role not in ['student', 'teacher', 'admin']:
        return error_response(msg='无效的角色类型', code=400)

    if User.objects.filter(username=username).exists():
        return error_response(msg='用户名已存在', code=400)

    new_user = User.objects.create_user(
        username=username,
        password=password,
        role=role,
        email=email,
        phone=phone,
        real_name=real_name
    )

    return created_response(data={
        'user_id': new_user.id,
        'username': new_user.username,
        'role': new_user.role
    }, msg='用户创建成功')


@api_view(['PUT'])
@permission_classes([IsAuthenticated, IsAdmin])
def admin_user_update(request, user_id):
    """
    更新用户信息（管理员）
    PUT /api/admin/users/{user_id}
    
    可更新字段：email, phone, real_name, role
    """
    user = request.user

    try:
        target_user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return error_response(msg='用户不存在', code=404)

    if target_user.id == user.id and 'role' in request.data:
        return error_response(msg='不能修改自己的角色', code=400)

    allowed_fields = ['email', 'phone', 'real_name', 'role']
    updated = []

    for field in allowed_fields:
        if field in request.data:
            if field == 'role' and request.data[field] not in ['student', 'teacher', 'admin']:
                return error_response(msg='无效的角色类型')
            setattr(target_user, field, request.data[field])
            updated.append(field)

    if updated:
        target_user.save()

    return success_response(data={
        'user_id': target_user.id,
        'updated_fields': updated
    }, msg='用户信息已更新')


@api_view(['DELETE'])
@permission_classes([IsAuthenticated, IsAdmin])
def admin_user_delete(request, user_id):
    """
    删除用户（管理员）
    DELETE /api/admin/users/{user_id}
    """
    user = request.user

    if user_id == user.id:
        return error_response(msg='不能删除自己')

    try:
        target_user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return error_response(msg='用户不存在', code=404)

    username = target_user.username
    target_user.delete()

    return success_response(msg=f'用户 {username} 已删除')


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsAdmin])
def admin_user_reset_password(request, user_id):
    """
    重置用户密码（管理员）
    POST /api/admin/users/{user_id}/reset-password
    
    请求参数：
    - new_password: 新密码（选填，不提供则自动生成）
    """
    try:
        target_user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return error_response(msg='用户不存在', code=404)

    provided_password = request.data.get('new_password')

    if not provided_password:
        alphabet = string.ascii_letters + string.digits
        new_password = ''.join(secrets.choice(alphabet) for _ in range(12))
        auto_generated = True
    else:
        new_password = provided_password
        auto_generated = False

    target_user.set_password(new_password)
    target_user.save()

    response_data = {'user_id': target_user.id}
    if auto_generated:
        response_data['new_password'] = new_password

    return success_response(data=response_data, msg='密码已重置')


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsAdmin])
def admin_user_disable(request, user_id):
    """
    禁用用户（管理员）
    POST /api/admin/users/{user_id}/disable
    """
    user = request.user

    if user_id == user.id:
        return error_response(msg='不能禁用自己')

    try:
        target_user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return error_response(msg='用户不存在', code=404)

    target_user.is_active = False
    target_user.save()

    return success_response(msg=f'用户 {target_user.username} 已禁用')


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsAdmin])
def admin_user_enable(request, user_id):
    """
    启用用户（管理员）
    POST /api/admin/users/{user_id}/enable
    """
    try:
        target_user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return error_response(msg='用户不存在', code=404)

    target_user.is_active = True
    target_user.save()

    return success_response(msg=f'用户 {target_user.username} 已启用')


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsAdmin])
def admin_user_batch_delete(request):
    """
    批量删除用户（管理员）
    POST /api/admin/users/batch-delete

    请求参数：
    - user_ids: 用户ID列表
    """

    user_ids = request.data.get('user_ids', [])
    if not user_ids:
        return error_response(msg='请提供要删除的用户ID列表')

    user_ids = [uid for uid in user_ids if uid != request.user.id]

    deleted_count, _ = User.objects.filter(id__in=user_ids).delete()
    return success_response(data={'deleted_count': deleted_count}, msg=f'已删除 {deleted_count} 个用户')


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsAdmin])
def admin_user_import(request):
    """
    从 Excel/CSV 批量导入用户（管理员）
    POST /api/admin/users/import

    请求参数：
    - file: Excel 或 CSV 文件
    """

    uploaded_file = request.FILES.get('file')
    if not uploaded_file:
        return error_response(msg='请上传文件')

    filename = uploaded_file.name.lower()
    rows = []

    try:
        if filename.endswith('.csv'):
            decoded = uploaded_file.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(decoded))
            rows = list(reader)
        elif filename.endswith(('.xlsx', '.xls')):
            try:
                import openpyxl
            except ImportError:
                return error_response(msg='服务器未安装 openpyxl，无法处理 Excel 文件')

            wb = openpyxl.load_workbook(uploaded_file, read_only=True)
            ws = wb.active
            if ws is None:
                return error_response(msg='Excel 文件缺少活动工作表')
            headers = [str(cell.value or '').strip() for cell in next(ws.iter_rows(max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, [str(v or '').strip() for v in row])))
        else:
            return error_response(msg='仅支持 .csv / .xlsx 文件')
    except Exception as e:
        return error_response(msg=f'文件解析失败: {e}')

    created_count = 0
    skipped = []

    with transaction.atomic():
        for i, row in enumerate(rows, start=2):
            username = row.get('username') or row.get('用户名', '')
            password = row.get('password') or row.get('密码', '') or 'Edu@12345'
            role = row.get('role') or row.get('角色', 'student')
            email = row.get('email') or row.get('邮箱', '')
            real_name = row.get('real_name') or row.get('姓名', '')
            student_id = row.get('student_id') or row.get('学号', '')

            if not username:
                skipped.append(f'第{i}行：缺少用户名')
                continue
            if User.objects.filter(username=username).exists():
                skipped.append(f'第{i}行：用户名 {username} 已存在')
                continue

            u = User(username=username, role=role, email=email, real_name=real_name, student_id=student_id)
            u.set_password(password)
            u.save()
            created_count += 1

    return success_response(data={
        'created_count': created_count,
        'skipped': skipped[:50],
    }, msg=f'成功导入 {created_count} 个用户')


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdmin])
def admin_user_export(request):
    """
    导出用户列表为 CSV
    GET /api/admin/users/export
    """

    role = request.query_params.get('role')
    queryset = User.objects.all()
    if role:
        queryset = queryset.filter(role=role)

    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="users.csv"'
    response.write(UTF8_BOM)

    writer = csv.writer(response)
    writer.writerow(['ID', '用户名', '姓名', '学号/工号', '角色', '邮箱', '手机', '状态', '注册时间'])

    for u in queryset[:10000]:
        writer.writerow([
            u.id, u.username, u.real_name or '', u.student_id or '', u.role,
            u.email or '', u.phone or '',
            '正常' if u.is_active else '禁用',
            u.date_joined.strftime('%Y-%m-%d %H:%M:%S') if u.date_joined else '',
        ])

    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdmin])
def admin_user_template(request):
    """
    获取用户导入模板
    GET /api/admin/users/template
    """
    logger.debug('管理员下载用户导入模板: admin=%s', getattr(request.user, 'id', None))

    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="user_import_template.csv"'
    response.write(UTF8_BOM)

    writer = csv.writer(response)
    writer.writerow(['username', 'password', 'role', 'email', 'real_name', 'student_id'])
    writer.writerow(['zhangsan', 'Edu@12345', 'student', 'zhangsan@example.com', '张三', '2024001'])
    writer.writerow(['teacher1', 'Edu@12345', 'teacher', 'teacher1@example.com', '李老师', 'T001'])

    return response


# ============ 学生画像管理 ============

@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdmin])
def admin_get_all_student_profiles(request):
    """
    管理员查看所有学生画像
    GET /api/admin/student-profiles
    
    查询参数：
    - course_id: 按课程筛选
    - user_id: 按用户筛选
    - page: 页码
    - page_size: 每页数量
    """
    course_id = request.query_params.get('course_id')
    user_id = request.query_params.get('user_id')
    page, page_size = _parse_pagination(request.query_params)

    if user_id:
        students = User.objects.filter(id=user_id, role='student')
    else:
        students = User.objects.filter(role='student')

    total = students.count()
    start = (page - 1) * page_size
    students = list(students[start:start + page_size])
    student_ids = [s.id for s in students]

    mastery_qs = KnowledgeMastery.objects.filter(user_id__in=student_ids)
    if course_id:
        mastery_qs = mastery_qs.filter(course_id=course_id)

    mastery_by_user = {}
    for m in mastery_qs:
        mastery_by_user.setdefault(m.user_id, []).append(float(m.mastery_rate))

    ability_qs = AbilityScore.objects.filter(user_id__in=student_ids)
    if course_id:
        ability_qs = ability_qs.filter(course_id=course_id)

    ability_by_user = {}
    for a in ability_qs:
        if a.user_id not in ability_by_user:
            ability_by_user[a.user_id] = a

    enrollment_counts = dict(
        Enrollment.objects.filter(user_id__in=student_ids)
        .values('user_id').annotate(cnt=models.Count('id')).values_list('user_id', 'cnt')
    )

    profiles = []
    for student in students:
        rates = mastery_by_user.get(student.id, [])
        avg_mastery = sum(rates) / len(rates) if rates else 0
        ability_score = ability_by_user.get(student.id)
        ability_scores = ability_score.scores if ability_score else None

        profiles.append({
            'user_id': student.id,
            'username': student.username,
            'real_name': student.real_name,
            'student_id': student.student_id,
            'average_mastery': round(avg_mastery, 2),
            'ability_scores': ability_scores if isinstance(ability_scores, dict) else {},
            'courses_enrolled': enrollment_counts.get(student.id, 0)
        })

    return success_response(data={
        'total': total,
        'page': page,
        'page_size': page_size,
        'profiles': profiles
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdmin])
def admin_student_profile_detail(request, student_id):
    """
    获取指定学生画像详情（管理员）
    GET /api/admin/student-profiles/{student_id}
    """
    try:
        student = User.objects.get(id=student_id, role='student')
    except User.DoesNotExist:
        return error_response(msg='学生不存在', code=404)

    course_id = request.query_params.get('course_id')

    mastery_qs = KnowledgeMastery.objects.filter(user=student).select_related('knowledge_point')
    if course_id:
        mastery_qs = mastery_qs.filter(course_id=course_id)

    mastery_list = [{
        'knowledge_point_id': m.knowledge_point_id,
        'knowledge_point_name': m.knowledge_point.name if m.knowledge_point else '',
        'mastery_rate': float(m.mastery_rate),
    } for m in mastery_qs]

    ability_qs = AbilityScore.objects.filter(user=student)
    if course_id:
        ability_qs = ability_qs.filter(course_id=course_id)
    ability = ability_qs.first()

    habit_pref = HabitPreference.objects.filter(user=student).first()

    from assessments.models import AnswerHistory
    total_answers = AnswerHistory.objects.filter(user=student).count()
    correct_answers = AnswerHistory.objects.filter(user=student, is_correct=True).count()

    return success_response(data={
        'user_id': student.id,
        'username': student.username,
        'real_name': student.real_name,
        'student_id': student.student_id,
        'knowledge_mastery': mastery_list,
        'ability_scores': ability.scores if ability else {},
        'habit_preferences': HabitPreferenceSerializer(habit_pref).data if habit_pref else {},
        'answer_stats': {
            'total': total_answers,
            'correct': correct_answers,
            'accuracy': round(correct_answers / total_answers * 100, 1) if total_answers > 0 else 0
        },
    })
