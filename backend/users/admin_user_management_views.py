"""用户模块 - 管理员用户管理接口。"""

import csv
import io
import logging
import secrets
import string

from django.db import models, transaction
from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated

from common.permissions import IsAdmin
from common.responses import created_response, error_response, success_response
from .admin_helpers import UTF8_BOM, _parse_pagination
from .models import User


logger = logging.getLogger(__name__)


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdmin])
def admin_user_list(request):
    """获取用户列表（管理员）。"""
    page, size = _parse_pagination(request.query_params, size_key='size')
    role = request.query_params.get('role')
    query = request.query_params.get('query')
    status = request.query_params.get('status')
    users = User.objects.all().order_by('-date_joined')

    if role:
        users = users.filter(role=role)
    if query:
        users = users.filter(
            models.Q(username__icontains=query)
            | models.Q(email__icontains=query)
            | models.Q(phone__icontains=query)
            | models.Q(real_name__icontains=query)
        )
    if status == 'active':
        users = users.filter(is_active=True)
    elif status == 'inactive':
        users = users.filter(is_active=False)

    total = users.count()
    users = users[(page - 1) * size : page * size]
    return success_response(data={
        'total': total,
        'page': page,
        'size': size,
        'users': [
            {
                'user_id': user.id,
                'username': user.username,
                'email': user.email or '',
                'phone': user.phone or '',
                'real_name': user.real_name or '',
                'role': user.role,
                'is_active': user.is_active,
                'date_joined': user.date_joined.isoformat(),
            }
            for user in users
        ],
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdmin])
def admin_user_detail(request, user_id):
    """获取用户详情（管理员）。"""
    try:
        target_user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return error_response(msg='用户不存在', code=404)

    return success_response(data={
        'user_id': target_user.id,
        'username': target_user.username,
        'email': target_user.email or '',
        'phone': target_user.phone or '',
        'real_name': target_user.real_name or '',
        'student_id': target_user.student_id or '',
        'role': target_user.role,
        'is_active': target_user.is_active,
        'avatar': target_user.avatar.url if target_user.avatar else None,
        'date_joined': target_user.date_joined.isoformat(),
        'last_login': target_user.last_login.isoformat() if target_user.last_login else None,
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsAdmin])
def admin_user_create(request):
    """创建用户（管理员）。"""
    username = request.data.get('username')
    password = request.data.get('password')
    role = request.data.get('role', 'student')
    if not username or not password:
        return error_response(msg='用户名和密码不能为空', code=400)
    if role not in ['student', 'teacher', 'admin']:
        return error_response(msg='无效的角色类型', code=400)
    if User.objects.filter(username=username).exists():
        return error_response(msg='用户名已存在', code=400)

    new_user = User.objects.create_user(
        username=username,
        password=password,
        role=role,
        email=request.data.get('email', ''),
        phone=request.data.get('phone', ''),
        real_name=request.data.get('real_name', ''),
    )
    return created_response(data={'user_id': new_user.id, 'username': new_user.username, 'role': new_user.role}, msg='用户创建成功')


@api_view(['PUT'])
@permission_classes([IsAuthenticated, IsAdmin])
def admin_user_update(request, user_id):
    """更新用户信息（管理员）。"""
    try:
        target_user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return error_response(msg='用户不存在', code=404)

    if target_user.id == request.user.id and 'role' in request.data:
        return error_response(msg='不能修改自己的角色', code=400)

    updated = []
    for field in ['email', 'phone', 'real_name', 'role']:
        if field not in request.data:
            continue
        if field == 'role' and request.data[field] not in ['student', 'teacher', 'admin']:
            return error_response(msg='无效的角色类型')
        setattr(target_user, field, request.data[field])
        updated.append(field)

    if updated:
        target_user.save()
    return success_response(data={'user_id': target_user.id, 'updated_fields': updated}, msg='用户信息已更新')


@api_view(['DELETE'])
@permission_classes([IsAuthenticated, IsAdmin])
def admin_user_delete(request, user_id):
    """删除用户（管理员）。"""
    if user_id == request.user.id:
        return error_response(msg='不能删除自己')
    try:
        target_user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return error_response(msg='用户不存在', code=404)

    username = target_user.username
    target_user.delete()
    return success_response(msg=f'用户 {username} 已删除')


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsAdmin])
def admin_user_reset_password(request, user_id):
    """重置用户密码（管理员）。"""
    try:
        target_user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return error_response(msg='用户不存在', code=404)

    provided_password = request.data.get('new_password')
    if not provided_password:
        alphabet = string.ascii_letters + string.digits
        new_password = ''.join(secrets.choice(alphabet) for _ in range(12))
        auto_generated = True
    else:
        new_password = provided_password
        auto_generated = False

    target_user.set_password(new_password)
    target_user.save()
    response_data = {'user_id': target_user.id}
    if auto_generated:
        response_data['new_password'] = new_password
    return success_response(data=response_data, msg='密码已重置')


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsAdmin])
def admin_user_disable(request, user_id):
    """禁用用户（管理员）。"""
    if user_id == request.user.id:
        return error_response(msg='不能禁用自己')
    try:
        target_user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return error_response(msg='用户不存在', code=404)
    target_user.is_active = False
    target_user.save()
    return success_response(msg=f'用户 {target_user.username} 已禁用')


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsAdmin])
def admin_user_enable(request, user_id):
    """启用用户（管理员）。"""
    try:
        target_user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return error_response(msg='用户不存在', code=404)
    target_user.is_active = True
    target_user.save()
    return success_response(msg=f'用户 {target_user.username} 已启用')


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsAdmin])
def admin_user_batch_delete(request):
    """批量删除用户（管理员）。"""
    user_ids = request.data.get('user_ids', [])
    if not user_ids:
        return error_response(msg='请提供要删除的用户ID列表')

    user_ids = [uid for uid in user_ids if uid != request.user.id]
    deleted_count, _ = User.objects.filter(id__in=user_ids).delete()
    return success_response(data={'deleted_count': deleted_count}, msg=f'已删除 {deleted_count} 个用户')


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsAdmin])
def admin_user_import(request):
    """从 Excel/CSV 批量导入用户（管理员）。"""
    uploaded_file = request.FILES.get('file')
    if not uploaded_file:
        return error_response(msg='请上传文件')

    filename = uploaded_file.name.lower()
    rows = []
    try:
        if filename.endswith('.csv'):
            decoded = uploaded_file.read().decode('utf-8-sig')
            rows = list(csv.DictReader(io.StringIO(decoded)))
        elif filename.endswith(('.xlsx', '.xls')):
            try:
                import openpyxl
            except ImportError:
                return error_response(msg='服务器未安装 openpyxl，无法处理 Excel 文件')
            workbook = openpyxl.load_workbook(uploaded_file, read_only=True)
            worksheet = workbook.active
            if worksheet is None:
                return error_response(msg='Excel 文件缺少活动工作表')
            headers = [str(cell.value or '').strip() for cell in next(worksheet.iter_rows(max_row=1))]
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, [str(value or '').strip() for value in row])))
        else:
            return error_response(msg='仅支持 .csv / .xlsx 文件')
    except Exception as exc:
        return error_response(msg=f'文件解析失败: {exc}')

    created_count = 0
    skipped = []
    with transaction.atomic():
        for index, row in enumerate(rows, start=2):
            username = row.get('username') or row.get('用户名', '')
            password = row.get('password') or row.get('密码', '') or 'Edu@12345'
            role = row.get('role') or row.get('角色', 'student')
            if not username:
                skipped.append(f'第{index}行：缺少用户名')
                continue
            if User.objects.filter(username=username).exists():
                skipped.append(f'第{index}行：用户名 {username} 已存在')
                continue
            user = User(
                username=username,
                role=role,
                email=row.get('email') or row.get('邮箱', ''),
                real_name=row.get('real_name') or row.get('姓名', ''),
                student_id=row.get('student_id') or row.get('学号', ''),
            )
            user.set_password(password)
            user.save()
            created_count += 1

    return success_response(data={'created_count': created_count, 'skipped': skipped[:50]}, msg=f'成功导入 {created_count} 个用户')


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdmin])
def admin_user_export(request):
    """导出用户列表为 CSV。"""
    role = request.query_params.get('role')
    queryset = User.objects.all()
    if role:
        queryset = queryset.filter(role=role)

    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="users.csv"'
    response.write(UTF8_BOM)
    writer = csv.writer(response)
    writer.writerow(['ID', '用户名', '姓名', '学号/工号', '角色', '邮箱', '手机', '状态', '注册时间'])
    for user in queryset[:10000]:
        writer.writerow([
            user.id,
            user.username,
            user.real_name or '',
            user.student_id or '',
            user.role,
            user.email or '',
            user.phone or '',
            '正常' if user.is_active else '禁用',
            user.date_joined.strftime('%Y-%m-%d %H:%M:%S') if user.date_joined else '',
        ])
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdmin])
def admin_user_template(request):
    """获取用户导入模板。"""
    logger.debug('管理员下载用户导入模板: admin=%s', getattr(request.user, 'id', None))
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="user_import_template.csv"'
    response.write(UTF8_BOM)
    writer = csv.writer(response)
    writer.writerow(['username', 'password', 'role', 'email', 'real_name', 'student_id'])
    writer.writerow(['zhangsan', 'Edu@12345', 'student', 'zhangsan@example.com', '张三', '2024001'])
    writer.writerow(['teacher1', 'Edu@12345', 'teacher', 'teacher1@example.com', '李老师', 'T001'])
    return response
