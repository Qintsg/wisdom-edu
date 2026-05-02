"""管理员用户管理接口的查询、导入导出与写入支撑逻辑。"""

from __future__ import annotations

import csv
import io
import secrets
import string
from collections.abc import Iterable, Mapping, Sequence
from typing import Protocol

from django.db import transaction
from django.db.models import Q, QuerySet
from django.http import HttpResponse

from .admin_helpers import UTF8_BOM
from .models import User


CSV_CONTENT_TYPE = "text/csv; charset=utf-8-sig"
DEFAULT_IMPORT_PASSWORD = "Edu@12345"
VALID_USER_ROLES = {"student", "teacher", "admin"}
USER_EXPORT_HEADER = ["ID", "用户名", "姓名", "学号/工号", "角色", "邮箱", "手机", "状态", "注册时间"]
USER_TEMPLATE_HEADER = ["username", "password", "role", "email", "real_name", "student_id"]
USER_TEMPLATE_ROWS = [
    ["zhangsan", DEFAULT_IMPORT_PASSWORD, "student", "zhangsan@example.com", "张三", "2024001"],
    ["teacher1", DEFAULT_IMPORT_PASSWORD, "teacher", "teacher1@example.com", "李老师", "T001"],
]


class UploadedUserFile(Protocol):
    """用户导入文件只依赖文件名和二进制读取能力。"""

    name: str

    def read(self) -> bytes:
        """读取上传文件内容。"""


def build_admin_user_list_payload(query_params: Mapping[str, object], page: int, size: int) -> dict[str, object]:
    """按筛选条件读取用户列表并组装分页响应。"""
    users = filtered_admin_users(query_params)
    total = users.count()
    page_users = users[(page - 1) * size : page * size]
    return {
        "total": total,
        "page": page,
        "size": size,
        "users": [build_admin_user_list_item(user) for user in page_users],
    }


def filtered_admin_users(query_params: Mapping[str, object]) -> QuerySet[User]:
    """构造管理员用户列表查询集，保留前端现有 query/size/status 参数契约。"""
    users = User.objects.all().order_by("-date_joined")
    role = query_params.get("role")
    query = query_params.get("query")
    status = query_params.get("status")

    if role:
        users = users.filter(role=role)
    if query:
        users = users.filter(
            Q(username__icontains=query)
            | Q(email__icontains=query)
            | Q(phone__icontains=query)
            | Q(real_name__icontains=query)
        )
    if status == "active":
        users = users.filter(is_active=True)
    elif status == "inactive":
        users = users.filter(is_active=False)
    return users


def build_admin_user_list_item(user: User) -> dict[str, object]:
    """序列化管理员用户列表项。"""
    return {
        "user_id": user.id,
        "username": user.username,
        "email": user.email or "",
        "phone": user.phone or "",
        "real_name": user.real_name or "",
        "role": user.role,
        "is_active": user.is_active,
        "date_joined": user.date_joined.isoformat(),
    }


def get_admin_user(user_id: int) -> User | None:
    """按主键读取用户，供多个管理端动作复用统一的不存在处理。"""
    return User.objects.filter(id=user_id).first()


def build_admin_user_detail_payload(target_user: User) -> dict[str, object]:
    """序列化管理员用户详情响应。"""
    return {
        "user_id": target_user.id,
        "username": target_user.username,
        "email": target_user.email or "",
        "phone": target_user.phone or "",
        "real_name": target_user.real_name or "",
        "student_id": target_user.student_id or "",
        "role": target_user.role,
        "is_active": target_user.is_active,
        "avatar": target_user.avatar.url if target_user.avatar else None,
        "date_joined": target_user.date_joined.isoformat(),
        "last_login": target_user.last_login.isoformat() if target_user.last_login else None,
    }


def create_admin_user(data: Mapping[str, object]) -> tuple[dict[str, object] | None, str | None]:
    """校验并创建管理员指定的新用户。"""
    username = data.get("username")
    password = data.get("password")
    role = data.get("role", "student")
    if not username or not password:
        return None, "用户名和密码不能为空"
    if role not in VALID_USER_ROLES:
        return None, "无效的角色类型"
    if User.objects.filter(username=username).exists():
        return None, "用户名已存在"

    new_user = User.objects.create_user(
        username=username,
        password=password,
        role=role,
        email=data.get("email", ""),
        phone=data.get("phone", ""),
        real_name=data.get("real_name", ""),
    )
    return {"user_id": new_user.id, "username": new_user.username, "role": new_user.role}, None


def update_admin_user(target_user: User, data: Mapping[str, object]) -> tuple[list[str] | None, str | None]:
    """按白名单字段更新用户，避免 View 层堆叠字段循环与校验。"""
    updated_fields: list[str] = []
    for field in ["email", "phone", "real_name", "role"]:
        if field not in data:
            continue
        field_value = data[field]
        if field == "role" and field_value not in VALID_USER_ROLES:
            return None, "无效的角色类型"
        setattr(target_user, field, field_value)
        updated_fields.append(field)

    if updated_fields:
        target_user.save()
    return updated_fields, None


def delete_admin_user(target_user: User) -> str:
    """删除用户并返回原用户名用于响应消息。"""
    username = target_user.username
    target_user.delete()
    return username


def reset_admin_user_password(target_user: User, provided_password: object) -> dict[str, object]:
    """重置密码；未显式传入时生成一次性随机密码并回传给管理员。"""
    auto_generated = not provided_password
    new_password = generate_random_password() if auto_generated else provided_password
    target_user.set_password(new_password)
    target_user.save()

    response_data: dict[str, object] = {"user_id": target_user.id}
    if auto_generated:
        response_data["new_password"] = new_password
    return response_data


def generate_random_password(length: int = 12) -> str:
    """生成仅含字母和数字的临时密码，兼容原有重置密码输出。"""
    alphabet = string.ascii_letters + string.digits
    return "".join(secrets.choice(alphabet) for _ in range(length))


def set_admin_user_active(target_user: User, is_active: bool) -> str:
    """切换用户启停状态并返回操作消息。"""
    target_user.is_active = is_active
    target_user.save()
    action = "启用" if is_active else "禁用"
    return f"用户 {target_user.username} 已{action}"


def delete_admin_users(raw_user_ids: object, excluded_user_id: int) -> int:
    """批量删除用户，显式排除当前管理员账号。"""
    user_ids = normalize_delete_user_ids(raw_user_ids, excluded_user_id)
    deleted_count, _ = User.objects.filter(id__in=user_ids).delete()
    return deleted_count


def normalize_delete_user_ids(raw_user_ids: object, excluded_user_id: int) -> list[object]:
    """将批量删除入参规整为列表，同时避免误删当前管理员。"""
    if isinstance(raw_user_ids, str):
        candidate_ids: Iterable[object] = [raw_user_ids]
    elif isinstance(raw_user_ids, Iterable):
        candidate_ids = raw_user_ids
    else:
        candidate_ids = []
    return [user_id for user_id in candidate_ids if user_id != excluded_user_id]


def read_user_import_rows(uploaded_file: UploadedUserFile) -> tuple[list[dict[str, str]], str | None]:
    """读取 CSV/XLS/XLSX 导入文件，返回标准行字典或用户可读错误。"""
    filename = uploaded_file.name.lower()
    try:
        if filename.endswith(".csv"):
            return read_user_csv_rows(uploaded_file), None
        if filename.endswith((".xlsx", ".xls")):
            return read_user_excel_rows(uploaded_file)
        return [], "仅支持 .csv / .xlsx 文件"
    except Exception as exc:
        return [], f"文件解析失败: {exc}"


def read_user_csv_rows(uploaded_file: UploadedUserFile) -> list[dict[str, str]]:
    """读取 UTF-8 BOM 兼容的 CSV 用户导入文件。"""
    decoded = uploaded_file.read().decode("utf-8-sig")
    return list(csv.DictReader(io.StringIO(decoded)))


def read_user_excel_rows(uploaded_file: UploadedUserFile) -> tuple[list[dict[str, str]], str | None]:
    """读取 Excel 用户导入文件，缺少 openpyxl 时返回明确部署错误。"""
    try:
        import openpyxl
    except ImportError:
        return [], "服务器未安装 openpyxl，无法处理 Excel 文件"

    workbook = openpyxl.load_workbook(uploaded_file, read_only=True)
    try:
        worksheet = workbook.active
        if worksheet is None:
            return [], "Excel 文件缺少活动工作表"
        headers = [str(cell.value or "").strip() for cell in next(worksheet.iter_rows(max_row=1))]
        rows = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            row_values = [str(value or "").strip() for value in row]
            rows.append(dict(zip(headers, row_values)))
        return rows, None
    finally:
        workbook.close()


def import_admin_users(rows: Sequence[Mapping[str, str]]) -> tuple[int, list[str]]:
    """批量创建导入用户，保持原有整批事务语义和最多 50 条跳过回传。"""
    created_count = 0
    skipped: list[str] = []
    with transaction.atomic():
        for index, row in enumerate(rows, start=2):
            was_created, skip_reason = create_user_from_import_row(row, index)
            if was_created:
                created_count += 1
            elif skip_reason:
                skipped.append(skip_reason)
    return created_count, skipped[:50]


def create_user_from_import_row(row: Mapping[str, str], index: int) -> tuple[bool, str | None]:
    """根据单行导入数据创建用户，兼容中英文字段名。"""
    username = first_row_value(row, "username", "用户名")
    password = first_row_value(row, "password", "密码") or DEFAULT_IMPORT_PASSWORD
    role = first_row_value(row, "role", "角色") or "student"
    if not username:
        return False, f"第{index}行：缺少用户名"
    if User.objects.filter(username=username).exists():
        return False, f"第{index}行：用户名 {username} 已存在"

    user = User(
        username=username,
        role=role,
        email=first_row_value(row, "email", "邮箱"),
        real_name=first_row_value(row, "real_name", "姓名"),
        student_id=first_row_value(row, "student_id", "学号"),
    )
    user.set_password(password)
    user.save()
    return True, None


def first_row_value(row: Mapping[str, str], *keys: str) -> str:
    """按字段别名读取导入行中的第一个非空值。"""
    for key in keys:
        value = row.get(key, "")
        if value:
            return value
    return ""


def build_user_export_response(query_params: Mapping[str, object]) -> HttpResponse:
    """导出管理员用户列表 CSV，限制最多 10000 行以控制响应体大小。"""
    role = query_params.get("role")
    queryset = User.objects.all()
    if role:
        queryset = queryset.filter(role=role)

    response = build_csv_response("users.csv")
    writer = csv.writer(response)
    writer.writerow(USER_EXPORT_HEADER)
    for user in queryset[:10000]:
        writer.writerow(build_user_export_row(user))
    return response


def build_user_export_row(user: User) -> list[object]:
    """组装用户导出 CSV 的单行内容。"""
    return [
        user.id,
        user.username,
        user.real_name or "",
        user.student_id or "",
        user.role,
        user.email or "",
        user.phone or "",
        "正常" if user.is_active else "禁用",
        user.date_joined.strftime("%Y-%m-%d %H:%M:%S") if user.date_joined else "",
    ]


def build_user_import_template_response() -> HttpResponse:
    """生成用户导入模板 CSV。"""
    response = build_csv_response("user_import_template.csv")
    writer = csv.writer(response)
    writer.writerow(USER_TEMPLATE_HEADER)
    writer.writerows(USER_TEMPLATE_ROWS)
    return response


def build_csv_response(filename: str) -> HttpResponse:
    """创建带 UTF-8 BOM 的 CSV 下载响应。"""
    response = HttpResponse(content_type=CSV_CONTENT_TYPE)
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    response.write(UTF8_BOM)
    return response
